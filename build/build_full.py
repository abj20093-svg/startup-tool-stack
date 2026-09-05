#!/usr/bin/env python3
"""The Startup Tool Stack: full 14-chapter self-contained HTML (Gate 3).
Chapter 1 copy is the Gate-1/Gate-2 approved text, verbatim.
Chapters 2-14 apply the identical procedure mechanically. Nothing is invented:
every fact is lifted from startup-tool-stack.xlsx.
"""
import html as H
import json
import math
import os
import re
import openpyxl

PROJ = "/Users/aryanbhardwaj/Downloads/Startup Decode Lead Magnet"
HERE = os.path.dirname(os.path.abspath(__file__))
VARIANT = os.environ.get("VARIANT") == "varied"
OUT = f"{HERE}/toolstack_varied.html" if VARIANT else f"{HERE}/toolstack_full.html"
VAR_MAP = {}
if VARIANT:
    for _c in json.load(open(f"{HERE}/variance_proposal.json")):
        VAR_MAP[_c["before"]] = _c["after"]

FONT_CSS = open(f"{HERE}/fonts/embedded.css").read()
TOOL_URLS = {k.lower(): v for k, v in json.load(open(f"{HERE}/tool_urls.json")).items()}
CSS_BODY = open(f"{HERE}/css.txt").read()

wb = openpyxl.load_workbook(f"{PROJ}/startup-tool-stack.xlsx", data_only=True)
src = wb[wb.sheetnames[0]]

# group, name, header_row, first_row, last_row, job label, decision line
CHAPTERS = [
    ("Selling", "CRM & Lead Gen", 3, 4, 16, "Manage leads and customers",
     "The decision: where your leads, deals and customer records live, and whether the same tool should also go find the leads."),
    ("Selling", "Email Sequencing", 18, 19, 30, "Send cold email at scale",
     "The decision: what sends your outbound email, and what keeps it out of spam folders."),
    ("Selling", "Newsletter / Marketing Campaigns", 32, 33, 45, "Run a newsletter",
     "The decision: where your subscriber list lives, what it costs as that list grows, and whether you can charge for it."),
    ("Selling", "Social Media Scheduling", 47, 48, 58, "Schedule social posts",
     "The decision: what queues your posts, which networks it can actually publish to on its own, and what it reports back."),
    ("Selling", "Payments & Billing", 60, 61, 81, "Take payments",
     "The decision: who processes your money, who is liable for the sales tax, and what you pay on each transaction."),
    ("Building", "App Building", 85, 86, 97, "Build your app fast",
     "The decision: what builds your first version, and whether you own the code that comes out."),
    ("Building", "Databases & Backend", 99, 100, 112, "Set up a database & backend",
     "The decision: where your data lives, what comes bundled with it, and what it costs you when traffic grows or stops."),
    ("Building", "Authentication", 114, 115, 126, "Add user login",
     "The decision: who handles sign-in and sessions, and what per-user pricing does to your bill as you grow."),
    ("Building", "Hosting", 128, 129, 141, "Host your app",
     "The decision: where your app runs, what it costs when traffic spikes, and how hard it is to leave."),
    ("Building", "Product Analytics & Experimentation", 143, 144, 155, "See how users use your product",
     "The decision: what records how people use your product, and whether the same tool runs your experiments and feature flags."),
    ("Operating", "Incorporation & Formation", 159, 160, 169, "Incorporate the company",
     "The decision: who files your company, what paperwork you end up holding, and what you keep paying each year."),
    ("Operating", "Bookkeeping & Accounting", 171, 172, 183, "Keep the books",
     "The decision: who keeps your books, whether they will stand up in a diligence review, and which accountants can work in them."),
    ("Operating", "Customer Support & Helpdesk", 185, 186, 198, "Support your customers",
     "The decision: where customer conversations land, which channels they arrive on, and what you pay per seat or per resolution."),
    ("Operating", "Workflow Automation", 200, 201, 211, "Automate workflows",
     "The decision: what connects your tools to each other, how it charges you for the work, and whether you can host it yourself."),
]

# ---- coverage guard: every content row must be declared ----
_declared = set()
for _g, _n, _h, _f, _l, _j, _d in CHAPTERS:
    _declared.add(_h)
    for _r in range(_f, _l + 1):
        assert _r not in _declared, f"row {_r} declared twice"
        _declared.add(_r)
_orphans = [r for r in range(1, 212) if r not in _declared and
            any(src.cell(row=r, column=c).value not in (None, "") for c in range(3, 10))]
assert not _orphans, f"uncovered source rows: {_orphans}"

NA_LOG, THIN_LOG, PAGE_SPANS, DEDUP_LOG = [], [], [], []

def esc(t):
    return H.escape(t, quote=False)

def fix_names(t):
    t = t.replace("Zoominfo", "ZoomInfo")
    # stage ranges read as words, not arrows (pricing-tier and workflow arrows kept)
    t = re.sub(r"(?i)\b(pre-seed)\s*→\s*(seed|series\s+a)\b", r"\1 to \2", t)
    return t


# Comparisons naming a same-chapter rival are rewritten to stand alone; the
# rival's own facts live in its own profile. Platform/ownership facts
# (Polar-on-Stripe, the Lemon Squeezy acquisition) are kept as information.
RIVAL_EDITS = [
 (", the longest public track record in this group (ZoomInfo, the other public company here, listed in 2020),", ", the longest public track record in this group,"),
 ("Cheaper tools (Apollo, Clay) suffice for early prospecting.", "Cheaper prospecting tools suffice at the early stage."),
 ("Salesforge/lemlist push harder on AI-written, per-prospect personalization at scale.", "AI-written, per-prospect personalization at scale is not the core bet here."),
 ("Steeper than SalesHandy for a first-timer.", "Steep for a first-timer."),
 (" (below Instantly's ~94%)", ", below the best result in this group"),
 ("not a personalization-at-scale engine like Salesforge", "not a personalization-at-scale engine"),
 ("AI features are lighter than Salesforge/Instantly; QuickMail bets", "AI features are light; QuickMail bets"),
 ("a shorter public track record than Instantly/QuickMail", "a shorter public track record"),
 ("depth of manual branching is lighter than Smartlead's API-driven flows", "depth of manual branching is lighter"),
 ("Ecosystem is younger/smaller than Smartlead's but coherent.", "Ecosystem is younger and smaller but coherent."),
 ("Smaller community and fewer playbooks than Instantly/Smartlead.", "Smaller community and fewer public playbooks."),
 ('less "unlimited-scale" than Instantly/Smartlead', 'not built for "unlimited-scale" volume'),
 ("Not as AI-heavy as Salesforge/Smartlead's fully AI-generated personalization at scale.", "Not AI-heavy; fully AI-generated personalization at scale is not the focus."),
 ("at high volume vs Instantly/Smartlead's flat unlimited sending", "at high volume"),
 ("Lighter marketing automation than Kit, and it", "Marketing automation is light, and it"),
 ("pricing below Mailchimp and Kit for comparable features", "pricing at the low end of this group for comparable features"),
 ("but no discovery/referral network like Substack, but you drive traffic yourself", "but there is no built-in discovery or referral network; you drive traffic yourself"),
 ("More design/branding control than Substack, but the interface", "Strong design and branding control, but the interface"),
 ("not deep reporting and is lighter than Metricool's analytics, a known tradeoff", "not deep reporting, a known tradeoff"),
 ("Analytics stay basic next to Metricool's, and", "Analytics stay basic, and"),
 ("No free plan (Hypefury also lacks one)", "No free plan"),
 (", but lighter than Metricool's reporting", ""),
 ("without the demographic, competitor and hashtag depth Metricool includes", "without deeper demographic, competitor or hashtag reporting"),
 ("Analytics are lighter than Metricool's, lacking the same demographic, competitor and hashtag depth.", "Analytics cover the basics, without deep demographic, competitor or hashtag reporting."),
 ("Fewer third-party integrations than Stripe.", "A smaller third-party integration ecosystem."),
 ("higher than assembling Stripe Billing plus separate processing", "higher than assembling billing tools plus separate processing yourself"),
 ("a shorter track record than Paddle's, and Stripe's holds", "a shorter track record, and Stripe's holds"),
 ("less no-code hand-holding than Lemon Squeezy or Gumroad", "little no-code hand-holding"),
 ("Subscription depth is narrower than Stripe's, and", "Subscription depth is narrow, and"),
 ("faster than Paddle's monthly cycle but slower and fee-bearing compared with a direct processor", "slower and fee-bearing compared with a direct processor"),
 ("Day-to-day holds are less reported than Paddle, yet", "Day-to-day holds are rarely reported, yet"),
 ("Integrations narrower than Stripe.", "The integration catalog is narrow."),
 ("or configurable dunning schedules like Stripe Billing", "or configurable dunning schedules"),
 ("Narrower global reach than Stripe or Adyen, mostly", "Narrow global reach, mostly"),
 ("well-documented but less deep than Stripe's developer tooling", "well-documented, though developer tooling is not its main strength"),
 ("No per-dispute fee (vs Stripe's $15), but", "No per-dispute fee, but"),
 ("Fewer third-party dev libraries and plugins than Stripe's ecosystem.", "A smaller ecosystem of third-party dev libraries and plugins."),
 ("Stripe has no comparable card-present + POS + hardware stack.", "No online-first processor pairs card-present, POS and first-party hardware this tightly."),
 ("Higher online rates than Stripe on the free plan.", "Online rates on the free plan are at the high end of this group."),
 (", though newer than Okta-owned Auth0,", ","),
 ("takes more effort than Clerk's drop-in approach", "takes more effort than a drop-in approach"),
 ("more configuration than a drop-in tool like Clerk", "more configuration than a drop-in tool"),
 ("it ships with less pre-built UI than Clerk.", "it ships with less pre-built UI."),
 ("Ships with less pre-built UI than Clerk, so you build", "Ships with less pre-built UI, so you build"),
 ("100K users cost roughly $1,000/mo on Clerk, a gap driven by the per-user rate.", "Per-user pricing elsewhere can reach roughly $1,000/mo at 100K users."),
 ("versus roughly $1,000/mo on Clerk, while keeping", "a fraction of per-user pricing at that scale, while keeping"),
 ("less turnkey/polished than Clerk's out-of-the-box components", "less turnkey than the most polished component kits"),
 ("Younger than Auth0 but mature and actively shipping.", "A younger vendor, but mature and actively shipping."),
 ("Drop-in UI is less polished than Clerk, and it has fewer legacy turnkey integrations/extensions than Auth0's older ecosystem.", "Drop-in UI is less polished, and it has fewer legacy turnkey integrations and extensions than older ecosystems."),
 ("Unlike Render, there is no ongoing free option.", "There is no ongoing free option."),
 ("Slower than Vercel.", "Slower than the fastest host here."),
 ("Serverless cold starts run ~3s+ (versus Vercel's ~1s), edge function cold starts ~28ms (vs ~12ms), and ~90ms average TTFB (vs ~70ms).", "Serverless cold starts run ~3s+, edge function cold starts ~28ms, and average TTFB ~90ms in 2026 benchmarks."),
 ("16+ edge locations versus Vercel's 100+, so distant users", "16+ edge locations, a comparatively small network, so distant users"),
 ("solid but fewer points of presence than Vercel (100+) or Cloudflare (300+), so users", "solid but a modest network by CDN standards, so users"),
 ("Surface price is close to Vercel's, but", "Surface price is in line with this group, but"),
 ("more predictable than Vercel's bandwidth/invocation overages", "more predictable than bandwidth- and invocation-metered billing"),
 ("it allows commercial use, unlike Vercel's Hobby tier", "its free tier allows commercial use"),
 ("Lower lock-in than Vercel.", "Lock-in is low."),
 ("allows commercial use (unlike Vercel), and background functions", "allows commercial use, and background functions"),
 ("Slower serverless cold starts (~3s) and far fewer edge locations (16+) than Vercel.", "Slower serverless cold starts (~3s) and a small edge network (16+ locations)."),
 ("so the latest Next.js features tend to land on Vercel first", "so the latest Next.js features tend to arrive late here"),
 ("as an alternative to Intercom, Zendesk and Salesforce Service Cloud for teams", "as a self-hosted alternative to the big hosted suites for teams"),
 ("Simpler tools (Intercom, Freshdesk, Help Scout) are cheaper", "Simpler tools are cheaper"),
 ("though it carries lighter enterprise-contract tooling than Paddle", "though enterprise-contract tooling is light"),
 ("Reach is built into the edge model, not something you provision.", "Reach is built into the edge model, not something you provision. The network spans 100+ edge locations; 2026 benchmarks measured ~1s serverless cold starts, ~12ms edge-function cold starts and ~70ms average TTFB."),
]
_RIVAL_HITS = set()
def scrub_comparisons(t):
    if not t:
        return t
    for _i, (_old, _new) in enumerate(RIVAL_EDITS):
        if _old in t:
            t = t.replace(_old, _new); _RIVAL_HITS.add(_i)
        else:
            _alt = _old.replace("'", "’")
            if _alt != _old and _alt in t:
                t = t.replace(_alt, _new.replace("'", "’")); _RIVAL_HITS.add(_i)
    return t

def cell(r, c):
    raw = str(src.cell(row=r, column=c).value or "").strip()
    if VARIANT and raw:
        raw = "\n".join(VAR_MAP.get(p, p) for p in raw.split("\n"))
    # keep prose references to the seventh column consistent with its label
    return scrub_comparisons(fix_names(raw).replace('Manual DIY', 'Internal build'))

def cell_raw(r, c):
    """Original workbook text, never varied: used for the At a glance table and the
    header price line, which the brief excludes from prose changes."""
    return scrub_comparisons(fix_names(str(src.cell(row=r, column=c).value or "").strip()))

def slug(name):
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")

ABBREV = ("incl", "e.g", "i.e", "vs", "approx", "est", "inc", "ltd", "co", "etc", "min",
          "max", "mo", "yr", "no", "st", "mr", "ms", "dr", "sept", "jan", "feb", "aug",
          "u.s", "a.m", "p.m", "fig", "ref", "cf")

def split_sents(t):
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9\"'$~(])", t.replace("\n", " ").strip())
    out = []
    for p in parts:
        if out:
            prev = out[-1]
            last_word = re.findall(r"[A-Za-z.]+", prev)
            unbalanced = (prev.count("(") != prev.count(")")
                          or prev.count("[") != prev.count("]"))
            is_abbrev = bool(last_word) and last_word[-1].rstrip(".").lower() in ABBREV
            if unbalanced or is_abbrev:
                out[-1] = prev + " " + p
                continue
        out.append(p)
    return out

LEADINS = ("Best when", "Avoid if", "Best fit", "Catch", "The catch", "Note")

def na_fix(t, where):
    out = t
    if re.match(r"^N/A as ", out):
        out = "Not applicable as " + out[len("N/A as "):]
        NA_LOG.append(where)
    elif re.match(r"^N/A:\s*", out):
        out = "Not applicable: " + re.sub(r"^N/A:\s*", "", out)
        NA_LOG.append(where)
    elif re.match(r"^N/A\b", out):
        out = "Not applicable" + out[3:]
        NA_LOG.append(where)
    return out

def strip_prefix(t):
    for p in ("Built for:", "Ideal for:"):
        if t.startswith(p):
            t = t[len(p):].strip()
            break
    return t[:1].upper() + t[1:] if t else t

def lead_split(t):
    m = re.match(r"^(.{15,90}?[.!?])\s+(.+)$", t, re.S)
    if m and not re.search(r"\d\.$", m.group(1)):
        return m.group(1).rstrip(".!?"), ".", m.group(2)
    m = re.match(r"^(.{15,80}?):\s+(.+)$", t, re.S)
    if m:
        return m.group(1), ":", m.group(2)
    m = re.match(r"^(.{25,90}?),\s+(.+)$", t, re.S)
    if m:
        return m.group(1), ",", m.group(2)
    m = re.match(r"^(.{15,130}?[.!?])\s+(.+)$", t, re.S)
    if m and not re.search(r"\d\.$", m.group(1)):
        return m.group(1).rstrip(".!?"), ".", m.group(2)
    m = re.match(r"^(.{15,110}?);\s+(.+)$", t, re.S)
    if m:
        return m.group(1), ";", m.group(2)
    return None, "", t

def smart_lead2(t):
    """(lead, delimiter, rest): verbatim lift; relaxed floor so short answers split too."""
    for pat, d in ((r"^(.{15,90}?)([.!?])\s+(.+)$", None), (r"^(.{12,80}?):\s+(.+)$", ":"),
                   (r"^(.{12,95}?)\s+(\(.+)$", ""), (r"^(.{18,95}?),\s+(.+)$", ","),
                   (r"^(.{15,110}?);\s+(.+)$", ";"), (r"^(.{15,130}?)([.!?])\s+(.+)$", None)):
        m = re.match(pat, t, re.S)
        if not m:
            continue
        if d is None:
            return m.group(1), m.group(2), m.group(3)
        if d == "":
            return m.group(1), "", m.group(2)
        return m.group(1), d, m.group(2)
    return None, "", t



def cap_first(t):
    return (t[:1].upper() + t[1:]) if t else t

def cap_if_new_sentence(detail, source_text):
    """Capitalise a detail line only when it starts a new sentence in the source.
    A detail that continues the lead's own sentence stays lower case."""
    if not detail:
        return detail
    flat = re.sub(r"\s+", " ", source_text)
    probe = re.sub(r"\s+", " ", detail)[:40]
    i = flat.find(probe)
    if i < 0:
        i = flat.lower().find(probe.lower())
    if i > 0:
        before = flat[:i].rstrip()
        if before.endswith((".", "!", "?")):
            return cap_first(detail)
        return detail
    return cap_first(detail) if re.match(r"^[A-Z][a-z]", probe) else detail

# ---------------- lead derivation (§4): a complete grammatical phrase ----------------
DANGLING = set("""and or but with without for to of in on at from by as than that which
while when because so plus including via per over under into onto a an the its their your
our this these those is are was were be been not no if only also both either neither""".split())

SETUP_OPENERS = ("rather than", "instead of", "because", "while", "although", "though",
                 "unlike", "beyond", "apart from", "aside from", "despite", "when",
                 "in addition to", "on top of", "if ", "as a", "for teams", "at entry")

MIDLEAD_LOG = []

def _balanced(s):
    return (s.count("(") == s.count(")") and s.count("[") == s.count("]")
            and s.count('"') % 2 == 0 and s.count("\u201c") == s.count("\u201d"))

def _ok_lead(s):
    """A lead must stand alone: balanced brackets, real substance, no dangling word."""
    s = s.strip().rstrip(",;:—- ")
    if not s or not _balanced(s):
        return False
    words = re.findall(r"[A-Za-z0-9$%/+.'-]+", s)
    if len(words) < 4:
        return False
    if words[-1].lower().strip(".") in DANGLING:
        return False
    return True

def derive_lead(text, where=""):
    """Return (lead, delimiter, rest). Lead is lifted verbatim and is grammatically whole."""
    t = re.sub(r"\s+", " ", text.strip())
    ss = split_sents(t)
    s1 = ss[0] if ss else t
    tail = " ".join(ss[1:]).strip()
    from_mid = False

    # if the sentence opens with setup, the claim is the main clause after it
    low = s1.lower()
    prefix = ""
    if any(low.startswith(op) for op in SETUP_OPENERS):
        for cm in re.finditer(r",\s+", s1):
            cand = s1[cm.end():].strip()
            first = re.findall(r"[A-Za-z'-]+", cand)
            if not first or first[0].lower().endswith("ly"):
                continue                      # adverb continuation, not the main clause
            if not _ok_lead(cand.rstrip(".!?")):
                continue
            prefix = s1[:cm.start() + 1].rstrip()
            s1 = cand
            from_mid = True
            break
    else:
        prefix = ""

    term = s1[-1] if s1 and s1[-1] in ".!?" else ""
    core = s1[:-1] if term else s1

    def finish(lead, delim, right):
        rest = (right + term).strip() if right else ""
        rest = (rest + (" " + tail if tail else "")).strip()
        if from_mid and where:
            MIDLEAD_LOG.append((where, prefix, lead))
        return (prefix + (" " if prefix else ""), lead, delim, rest)

    # a genuinely short answer is its own lead; longer ones lead with their opening claim
    if len(core) < 60 and _ok_lead(core):
        return finish(core, term or ".", "")

    best = None
    for m in re.finditer(r"\s+—\s+|:\s+|;\s+|,\s+(?=and |but |so |which |though |while |with )|,\s+", core):
        left, right = core[:m.start()].strip(), core[m.end():].strip()
        if len(left) < 28 or not _ok_lead(left):
            continue
        d = m.group(0).strip()
        d = d if d in (":", ";", ",") else ","
        best = (left, d, right)   # first complete phrase wins: a lead is the claim, not the list
        break
    if best:
        return finish(best[0], best[1], best[2])
    return finish(core, term or ".", "")   # no clean boundary: the whole claim leads

def para_html(p):
    for li in LEADINS:
        if p.startswith(li + ":"):
            return f'<p><span class="lead">{esc(li)}:</span>{esc(p[len(li)+1:])}</p>'
    return f"<p>{esc(p)}</p>"

def prose(t, where="", with_lead=True, force_lead=False):
    """force_lead: factor answers carry no sub-heading, so the bold lead is the
    only thing separating them, every one gets a lead, short answers included."""
    paras = [p.strip() for p in t.split("\n") if p.strip()]
    out = []
    for i, p in enumerate(paras):
        p = na_fix(p, where) if i == 0 else p
        if p.startswith("Not applicable"):
            m = re.match(r"^(Not applicable[^:—]*)\s*[:—]\s*(.*)$", p, re.S)
            if m:
                head, rest = m.group(1), m.group(2)
            else:
                head, rest = p, ""
            frag = f'<span class="na">{esc(head.strip())}</span>'
            rest = rest.strip()
            if rest:
                if force_lead:
                    pre2, l2, d2, r2 = derive_lead(rest, where)
                    frag += ": " + (esc(pre2) if pre2 else "")
                    frag += f'<strong class="pl">{esc(l2)}</strong>{d2}'
                    if r2:
                        frag += " " + esc(r2)
                else:
                    frag += ": " + esc(rest)
            out.append(f"<p>{frag}</p>")
            continue
        if with_lead and i == 0 and (force_lead or len(p) > 140):
            pre, lead, delim, rest = derive_lead(p, where)
            if lead:
                frag = esc(pre) if pre else ""
                frag += f'<strong class="pl">{esc(lead)}</strong>{delim}'
                if rest:
                    frag += " " + esc(rest)
                out.append(f"<p>{frag}</p>")
                continue
        out.append(f"<p>{esc(p)}</p>")
    return "".join(out)



def _split_semis(s):
    """Split on '; ' only at paren depth zero."""
    parts, depth, start = [], 0, 0
    i = 0
    while i < len(s):
        ch = s[i]
        if ch in "([":
            depth += 1
        elif ch in ")]":
            depth = max(0, depth - 1)
        elif ch == ";" and depth == 0 and i + 1 < len(s) and s[i + 1] == " ":
            parts.append(s[start:i])
            start = i + 2
            i += 1
        i += 1
    parts.append(s[start:])
    return parts

def _sub_items(rest):
    items = []
    for s in split_sents(rest or ""):
        for seg in _split_semis(s):
            seg = seg.strip().rstrip(";")
            if not seg:
                continue
            if seg[-1] not in ".!?":
                seg += "."
            items.append(cap_first(seg))
    return items

def bullets(t, where="", with_lead=True):
    """Factor text as a bulleted list: the lead claim is the bullet, the
    remaining sentences are its sub-bullets."""
    paras = [p.strip() for p in t.split("\n") if p.strip()]
    lis = []
    for i, p in enumerate(paras):
        p = na_fix(p, where) if i == 0 else p
        if p.startswith("Not applicable"):
            m = re.match(r"^(Not applicable[^:—]*)\s*[:—]\s*(.*)$", p, re.S)
            head, rest = (m.group(1), m.group(2)) if m else (p, "")
            head_html = f'<span class="na">{esc(head.strip())}</span>'
            subs = _sub_items(rest.strip())
        elif with_lead:
            pre, lead, delim, rest = derive_lead(p, where)
            lead_txt = lead.rstrip(".") + "."
            head_html = (esc(pre) if pre else "") + f'<strong class="pl">{esc(lead_txt)}</strong>'
            subs = _sub_items(rest)
        else:
            ss = split_sents(p)
            head_html = esc(ss[0]) if ss else esc(p)
            subs = _sub_items(" ".join(ss[1:]))
        sub_html = ("<ul>" + "".join(f"<li>{esc(s)}</li>" for s in subs) + "</ul>") if subs else ""
        lis.append(f"<li>{head_html}{sub_html}</li>")
    return f'<ul class="fb">{"".join(lis)}</ul>'

def parse_choose(t):
    parts = {"ideal": "", "best": "", "avoid": ""}
    for p in [q.strip() for q in t.split("\n") if q.strip()]:
        for key, pre in (("ideal", "Ideal for:"), ("best", "Best when:"), ("avoid", "Avoid if:")):
            if p.startswith(pre):
                parts[key] = p[len(pre):].strip()
    return parts


def dedup_against(text, seen, where):
    """§2: a sentence already printed earlier in this profile is not reprinted.
    The fact survives in its first position; nothing is lost."""
    out_paras = []
    for para in [p for p in text.split("\n") if p.strip()]:
        keep = []
        for sent in split_sents(para):
            key = re.sub(r"[^a-z0-9]", "", sent.lower())
            if len(key) > 40 and key in seen:
                DEDUP_LOG.append((where, sent[:90]))
                continue
            if len(key) > 40:
                seen.add(key)
            keep.append(sent)
        if keep:
            out_paras.append(" ".join(keep))
    return "\n".join(out_paras)

# ---------------- band classification (uniform, all 14 chapters) ----------------
def band_of(lab):
    L = lab.lower()
    if re.search(r"pricing models it handles|payment methods|legal support", L):
        return "What it does"
    if re.search(r"\b(pricing|price|fees?|costs?|free tier|overage|credit burn|chargeback|dispute)\b", L):
        return "Cost"
    if re.search(r"\b(integrations?|migration|lock-in|portability|self-hosting|exit|ownership|leave)\b", L):
        return "Ecosystem & exit"
    if re.search(r"\b(maturity|track record|support|trust|reliability)\b", L):
        return "Track record"
    return "What it does"

BAND_ORDER = ["What it does", "Cost", "Ecosystem & exit", "Track record"]

# ---------------- Gate-1 approved chapter 1 copy (verbatim, never re-derived) ----
CH1_GLANCE = {
 "HubSpot CRM": {
  "choose": ("Pre-seed to Series A B2B founders who want one system to grow into", "non-technical, US/EU, starting free (2 seats, 1,000 contacts)."),
  "price": ("Free", "2 seats, 1,000 contacts; then $20/user/mo, or $7 on annual billing."),
  "limit": ("Cost and complexity both climb with scale", "the free entry point is broad, but each hub added and each step up a tier raises the price and the number of moving parts, and capabilities are gated behind tiers rather than included."),
 },
 "Attio": {
  "choose": ("Technical, product-led founders who treat CRM as a flexible database", "pre-seed to seed, startups, remote, B2B, wanting custom objects."),
  "price": ("Free", "3 users, 50,000 records; Plus $35/user/mo annual ($44 monthly), Pro $79 ($99 monthly)."),
  "limit": ("Spend is harder to forecast", "than with flat per-seat pricing, because the model combines seats with usage credits and the bill moves with consumption as well as headcount."),
 },
 "Pipedrive": {
  "choose": ("Sales-led founders who live inside a deal pipeline", "non-technical, SMB, B2B, global, small AE teams wanting cheap visual pipeline."),
  "price": ("$14/user/mo", "Lite, billed annually; Growth $39, Premium $49, Ultimate $79."),
  "limit": ("LeadBooster and Smart Docs are add-ons, not included features", "so a team that needs them pays more than the advertised tier suggests; the limits follow directly from the focus."),
 },
 "Close": {
  "choose": ("Founders whose reps live on the phone", "inside-sales, SMB, B2B, US-centric; small high-velocity SDR/AE teams doing call + SMS + email."),
  "price": ("$9/user/mo", "Solo; Essentials $35, Growth $99, Scale $139."),
  "limit": ("The features that define Close are behind the paywall", "the Power Dialer requires Growth at $99, while the Predictive Dialer and call coaching sit on Scale at $139 only, so Solo and Essentials leave out the capability most buyers arrive for."),
 },
 "Apollo": {
  "choose": ("Outbound founders who want a contact database AND sequencing in one tool", "seed-stage, B2B, semi-technical, US/global, building an SDR motion from scratch."),
  "price": ("Free tier", "then Basic ~$49, Professional ~$79, Organization ~$119/user/mo; cheaper on annual billing."),
  "limit": ("Data is confined to Apollo's own 240M+ database", "instead of aggregated sources, so coverage is capped at what Apollo holds; mobile numbers and export credits are limited and are where Apollo upsells hardest."),
 },
 "ZoomInfo": {
  "choose": ("Well-funded Series A+ teams with a dedicated RevOps/SDR org", "B2B, mid-market/enterprise motion, US-heavy, budget for annual contracts."),
  "price": ("Quote-only", "reportedly ~$15,000/yr to start (3 seats, ~5,000 credits), climbing to $40,000+; median contract ~$32,000/yr per Vendr data."),
  "limit": ("Expensive, quote-gated, 3-seat minimum and add-on-heavy", "a poor fit pre-revenue; annual lock-in, aggressive renewals, thinner international/SMB data, and privacy/compliance overhead."),
 },
 "Internal build": {
  "choose": ("Technical founders with under ~100 target accounts", "earliest pre-seed, near-zero budget, hands-on, doing founder-led sales personally."),
  "price": ("No license fee", "cost is time: setup hours plus your AI plan and API/token burn for enrichment."),
  "limit": ("Won't scale past small volume without real eng effort", "you own all security, compliance, and dedupe risk; maintenance is a permanent tax you keep paying."),
 },
}
CH1_HEADER_PRICE = {
    "HubSpot CRM": ("Free", "2 seats, 1,000 contacts, 1 pipeline; then $20/user/mo, or $7 on annual billing."),
    "Attio": ("Free", "3 users, 50,000 records; Plus $35/user/mo annual ($44 monthly), Pro $79 ($99 monthly), both raised ~July 2026."),
    "Pipedrive": ("$14/user/mo", "Lite, billed annually; Growth $39, Premium $49, Ultimate $79 (July 2025 rebrand of Essential/Advanced/Professional)."),
    "Close": ("$9/user/mo", "Solo; Essentials $35, Growth $99, Scale $139."),
    "Apollo": ("Free tier", "then Basic ~$49, Professional ~$79, Organization ~$119/user/mo; cheaper on annual billing."),
    "ZoomInfo": ("Quote-only", "reportedly ~$15,000/yr to start (3 seats, ~5,000 credits), climbing to $40,000+; median contract ~$32,000/yr per Vendr data."),
    "Internal build": ("No license fee", "cost is time: mostly one-time setup hours plus your AI plan and API/token burn for enrichment."),
}
CH1_PRICING_DROP = {"HubSpot CRM": [1, 2], "Attio": [1], "Pipedrive": [1], "Close": [1],
                    "Apollo": [1], "ZoomInfo": [3], "Internal build": [2]}

# ---------------- mechanical derivations for chapters 2-14 ----------------

def smart_lead(t):
    """Widened lift cascade for glance cells: always returns (lead, detail),
    lead lifted verbatim from t. Never invents, never truncates mid-word."""
    t = t.strip()
    pats = [
        r"^(.{15,90}?[.!?])\s+(.+)$",       # first sentence
        r"^(.{15,80}?):\s+(.+)$",           # colon head
        r"^(.{12,95}?)\s+(\(.+)$",          # before a parenthetical
        r"^(.{25,95}?),\s+(.+)$",           # first comma clause
        r"^(.{15,130}?[.!?])\s+(.+)$",      # longer sentence
        r"^(.{15,110}?);\s+(.+)$",          # semicolon clause
        r"^(.{25,120}?),\s+(.+)$",          # wider comma
        r"^(.{40,130}?)\s+(?:and|but|though|while|so|because)\s+(.+)$",
    ]
    for p in pats:
        m = re.match(p, t, re.S)
        if m and not re.search(r"\d\.$", m.group(1)):
            return m.group(1).rstrip(".,;:"), m.group(2)
    return t.rstrip("."), ""

def derive_price(pricing_text):
    """Lead = a complete, bracket-balanced figure phrase lifted from sentence 1."""
    ss = split_sents(pricing_text)
    s1 = ss[0] if ss else ""
    core = s1[:-1] if s1[-1:] in ".!?" else s1

    def price_ok(x):
        x = x.strip().rstrip(",;:")
        if not x or not _balanced(x):
            return False
        w = re.findall(r"[A-Za-z0-9$%/+.'-]+", x)
        return bool(w) and w[-1].lower().strip(".") not in DANGLING

    best = None
    for m in re.finditer(r"\s*[;:,]\s+|\s+(?=\()", core):
        left, right = core[:m.start()].strip(), core[m.end():].strip()
        if 8 <= len(left) <= 95 and price_ok(left):
            best = (left, right)
            if len(left) >= 20:
                break
    if best:
        return best[0].rstrip(".,;:"), best[1].strip(), len(ss)
    return core.rstrip(".,;:"), "", len(ss)

def derive_choose(ideal_text):
    t = strip_prefix(ideal_text)
    m = re.match(r"^(.{20,110}?):\s+(.+)$", t)
    if m:
        return m.group(1).rstrip(",;"), m.group(2)
    return smart_lead(t)

def derive_benefit(text):
    """Bold lead plus detail for the At a glance benefit column.

    The detail renders as its own block, so the lead has to end as a complete
    sentence. Splitting at a comma (as the limitations column did) leaves the
    bold line trailing off into white space."""
    first = text.split("\n")[0].strip()
    sents = [x.strip() for x in split_sents(first) if x.strip()]
    if not sents:
        return (first, "")
    lead, rest = sents[0], " ".join(sents[1:]).strip()
    if len(lead.split()) > 34 and rest == "":      # one very long sentence: keep it whole
        return (lead, "")
    return (lead, rest)


def derive_limit(limit_text):
    pre, lead, _d, rest = derive_lead(limit_text.split("\n")[0].strip())
    return ((pre + lead).strip(), rest)


def scrub_rivals(text, self_tool, tools):
    """Drop competitor mentions from the pitch sections. Rival names
    usually arrive as parenthetical asides, so those are excised first,
    keeping the sentence they decorate; only a clause that still names
    a rival after that is dropped whole."""
    if not text:
        return text
    rivals = [t for t in tools if t != self_tool]
    pat = re.compile(r"\b(" + "|".join(re.escape(t) for t in rivals) + r")\b")
    text = re.sub(r"\s*\(([^()]*)\)",
                  lambda m: "" if pat.search(m.group(1)) else m.group(0), text)
    parts = re.split(r"(?<=[.!?;])\s+", text.strip())
    kept = [p for p in parts if not pat.search(p)]
    out = " ".join(kept)
    out = re.sub(r"\s+,", ",", re.sub(r"\s{2,}", " ", out)).strip().rstrip(";,")
    if out and out[-1] not in ".!?":
        out += "."
    return out


def xlink_avoid(text_html, self_tool, tools, cslug):
    out = text_html
    for other in sorted(tools, key=len, reverse=True):
        if other == self_tool:
            continue
        out = re.sub(r"(?<![\w>/-])" + re.escape(other) + r"(?![\w<-])",
                     f'<a class="xlink" href="#tool-{cslug}-{slug(other)}">{other}</a>', out)
    return out

css = FONT_CSS + CSS_BODY
PF = ('<div class="pfoot" aria-hidden="true">The Startup Tool Stack · '
      'Last updated August 2026</div>')

# ================================ build chapters ================================
chapter_html = []
contents_rows = {"Selling": [], "Building": [], "Operating": []}
jobs_html = ""
WIZ = []
all_tool_ids = set()

chapter_nav = []
for ci, (group, name, hrow, frow, lrow, job, decision) in enumerate(CHAPTERS, start=1):
    WIZ_TOOLS = []
    is_ch1 = (ci == 1)
    cslug = slug(name)
    tools = [cell(hrow, c) for c in range(3, 10)]
    # The seventh column is the do-it-internally option. Renamed for the audience;
    # the workbook keeps its original wording.
    tools = ["Internal build" if t.strip().lower() == "manual diy" else t for t in tools]
    labels = {r: fix_names(re.sub(r"^\d+\.\s*", "", str(src.cell(row=r, column=2).value)))
              for r in range(frow, lrow + 1)}

    def row_for(pred):
        for r, lb in labels.items():
            if pred(lb):
                return r
        return None

    r_built = row_for(lambda lb: lb == "Built for")
    r_what = row_for(lambda lb: lb == "What it is")
    r_stands = row_for(lambda lb: lb == "Where it stands out")
    r_limits = row_for(lambda lb: lb == "Limitations")
    r_choose = row_for(lambda lb: lb.startswith("Choose this"))
    special = {r_built, r_what, r_stands, r_limits, r_choose}
    cap_rows = [r for r in range(frow, lrow + 1) if r not in special]
    banded = {b: [r for r in cap_rows if band_of(labels[r]) == b] for b in BAND_ORDER}
    r_price = banded["Cost"][0] if banded["Cost"] else None

    # thin-source logging
    for r in range(frow, lrow + 1):
        for i, c in enumerate(range(3, 10)):
            t = cell(r, c)
            if 0 < len(t) < 80:
                THIN_LOG.append((name, tools[i], labels[r], len(t)))

    # ---- glance rows + header prices ----
    glance, header_price, price_drop = {}, {}, {}
    for i, c in enumerate(range(3, 10)):
        tool = tools[i]
        if is_ch1:
            _g = CH1_GLANCE[tool]
            glance[tool] = {
                "choose": (_g["choose"][0], cap_if_new_sentence(_g["choose"][1], cell_raw(r_choose, c))),
                "price": (_g["price"][0], cap_if_new_sentence(_g["price"][1], cell_raw(r_price, c) if r_price else "")),
                "limit": derive_benefit(cell(r_stands, c)),
            }
            header_price[tool] = CH1_HEADER_PRICE[tool]
            price_drop[tool] = []
            continue
        ptxt = cell_raw(r_price, c) if r_price else ""
        plead, pdetail, nsent = derive_price(ptxt)
        header_price[tool] = (plead, pdetail)
        _ch = derive_choose(parse_choose(cell_raw(r_choose, c))["ideal"])
        _lm = derive_benefit(cell(r_stands, c))
        glance[tool] = {
            "choose": (_ch[0], cap_if_new_sentence(_ch[1], cell_raw(r_choose, c))),
            "price": (plead, cap_if_new_sentence(pdetail, ptxt)),
            "limit": (_lm[0], _lm[1]),
        }
        # drop sentence 1 from Cost band only when other sentences remain
        price_drop[tool] = []

    def glance_cell(pair, attrs=""):
        lead, detail = pair
        d = f'<span class="gdetail">{esc(detail)}</span>' if detail else ""
        return f'<td {attrs}><span class="glead">{esc(lead)}</span>{d}</td>'

    grows = ""
    for tool in tools:
        g = glance[tool]
        _tu = TOOL_URLS.get(tool.lower())
        _tn = (f'<a class="site" href="{_tu}" target="_blank" rel="noopener noreferrer">{esc(tool)}</a>'
               if _tu else esc(tool))
        grows += (f'<tr><td class="tname">{_tn}</td>'
                  + glance_cell(g["choose"], 'data-label="Choose this if you are"')
                  + glance_cell(g["price"], 'data-label="Starting price" class="num"')
                  + glance_cell(g["limit"], 'data-label="Main benefit"') + "</tr>")

    # ---- profiles ----
    profiles = ""
    tool_anchors = []
    for i, c in enumerate(range(3, 10)):
        tool = tools[i]
        tid = f"tool-{cslug}-{slug(tool)}"
        tool_anchors.append((tool, tid))
        all_tool_ids.add(tid)
        diy = tool.strip().lower() in ("manual diy", "self-managed", "build", "internal build")
        what = cell(r_what, c)
        what_sents = split_sents(what)
        identity = what_sents[0].rstrip(".") if what_sents else ""
        # keep every remainder sentence, minus any that another factor already carries
        _others = re.sub(r"[^a-z0-9]", "", " ".join(
            cell(r, c) for r in range(frow, lrow + 1) if r != r_what).lower())
        what_rest = " ".join(x for x in what_sents[1:]
                             if re.sub(r"[^a-z0-9]", "", x.lower()) not in _others).strip()
        plead, pdetail = header_price[tool]
        choose = parse_choose(cell(r_choose, c))
        for _k in ("ideal", "best", "avoid"):
            choose[_k] = scrub_rivals(choose[_k], tool, tools)
        built = scrub_rivals(strip_prefix(cell(r_built, c)), tool, tools)

        CPL, LH = 49, 15
        def para_est(t, cpl=CPL):
            tot = 0
            for p in t.split("\n"):
                p = p.strip()
                if not p:
                    continue
                for k, s in enumerate(split_sents(p)):
                    eff = cpl if k == 0 else max(cpl - 5, 20)
                    tot += math.ceil(len(s) / eff) * LH + 7
                tot += 4
            return tot

        seen_sents = set()
        for _pre in (identity, built, choose["ideal"], choose["best"], choose["avoid"]):
            for _s in split_sents(_pre or ""):
                _k = re.sub(r"[^a-z0-9]", "", _s.lower())
                if len(_k) > 40: seen_sents.add(_k)
        units = []
        for band in BAND_ORDER:
            rows_in = banded[band]
            if band == "What it does" and what_rest:
                units.append((16 + 22 + para_est(what_rest),
                              f'<div class="bandlabel">{esc(band)}</div>'
                              f'<div class="factor">'
                              f'{bullets(dedup_against(what_rest, seen_sents, f"{name}/{tool}/What it is"), where=f"{name}/{tool}/What it is")}</div>'))
            for j, r in enumerate(rows_in):
                txt = cell(r, c)
                if r == r_price and price_drop[tool]:
                    ss = split_sents(txt)
                    txt = " ".join(s for k, s in enumerate(ss, 1) if k not in price_drop[tool])
                txt = dedup_against(txt, seen_sents, f"{name}/{tool}/{labels[r]}")
                body = bullets(txt, where=f"{name}/{tool}/{labels[r]}")
                fh = f'<div class="factor">{body}</div>'
                est = 16 + para_est(txt)
                if j == 0 and not (band == "What it does" and what_rest):
                    fh = f'<div class="bandlabel">{esc(band)}</div>' + fh
                    est += 22
                units.append((est, fh))
        for label, r in (("Where it stands out", r_stands), ("Limitations", r_limits)):
            txt = dedup_against(cell(r, c), seen_sents, f"{name}/{tool}/{label}")
            units.append((38 + para_est(txt),
                          f'<div class="bandlabel">{esc(label)}</div>'
                          f'<div class="factor">{bullets(txt, where=f"{name}/{tool}/{label}")}</div>'))

        h1 = 14 + 30 + math.ceil(len(identity) / 70) * LH + 10 + LH + 10
        h1 += 24 + para_est(built, 70) + 8
        ideal_t = strip_prefix(choose["ideal"])
        _persona = split_sents(ideal_t)[0] if ideal_t else tool
        WIZ_TOOLS.append([tool, tid, cap_first(_persona)])
        h1 += 34 + math.ceil(len(ideal_t) / 70) * LH
        h1 += 22 + max(math.ceil(len(choose["best"]) / 45),
                       math.ceil(len(choose["avoid"]) / 45)) * LH + 22
        cap1 = 700 - h1 - 36          # per-column budget, page 1 (after header block)
        capN = 640 - 34                # per-column budget, continuation pages
        pages_units, cur, col, acc = [], [[], []], 0, 0
        cap = cap1
        for est, uh in units:
            if acc + est > cap and cap > 60:
                if col == 0:
                    col, acc = 1, 0
                else:
                    pages_units.append(cur)
                    cur, col, acc, cap = [[], []], 0, 0, capN
            cur[col].append(uh); acc += est
        pages_units.append(cur)
        crumb_label = (f"the build route · {i+1} of {len(tools)}" if diy
                       else f"tool {i+1} of {len(tools)}")
        price_line = esc(plead)
        avoid_html = xlink_avoid(esc(cap_first(choose["avoid"])), tool, tools, cslug)
        cont_pages = ""
        for pi, (ca, cb) in enumerate(pages_units[1:], start=2):
            cont_pages += (f'<div class="ppage p2"><div class="runhead"><b>{esc(tool)}</b> &nbsp;·&nbsp; continued '
                           f'&nbsp;·&nbsp; {esc(name)}</div>'
                           f'<div class="bands b2"><div class="tcol">{"".join(ca)}</div>'
                           f'<div class="tcol">{"".join(cb)}</div></div>{PF}</div>')
        p1c1, p1c2 = pages_units[0]
        n_pages = len(pages_units)
        PAGE_SPANS.append((name, tool, n_pages))

        profiles += f"""
<article class="profile" id="{tid}" data-name="{esc(tool.lower())}" data-cat="{esc((name + ' ' + job).lower())}" data-chapter="{cslug}">
  <div class="ppage p1">
  <div class="phead">
    <h4>{esc(tool)}</h4>
  </div>
  <div class="psec">
    <div class="pseclabel">Built for</div>
    {bullets(built, where=f"{name}/{tool}/Built for")}
  </div>
  <div class="ledger">
    <div class="lhead">Choose this if you are</div>
    <div class="ideal">{bullets(ideal_t, where=f"{name}/{tool}/Ideal", with_lead=False)}</div>
    <div class="pair">
      {f'<div class="pcell go"><div class="plabel">Best when</div>{bullets(cap_first(choose["best"]), where=f"{name}/{tool}/Best", with_lead=False)}</div>' if choose["best"].strip() else ''}
      {f'<div class="pcell stop"><div class="plabel">Avoid if</div>{bullets(cap_first(choose["avoid"]), where=f"{name}/{tool}/Avoid", with_lead=False)}</div>' if choose["avoid"].strip() else ''}
    </div>
  </div>
  <div class="bands"><div class="tcol">{"".join(p1c1)}</div><div class="tcol">{"".join(p1c2)}</div></div>{PF}</div>
  {cont_pages}
</article>"""

    chips = ' &middot; '.join(f'<a class="tlink" href="#{a}">{esc(t)}</a>' for t, a in tool_anchors)
    WIZ.append({"s": cslug, "name": name, "job": job, "tools": WIZ_TOOLS})
    chapter_html.append(f"""
<section class="chapter" id="{cslug}" data-chapter="{cslug}">
  <div class="chap-open wrap">
    <h2>{esc(name)}</h2>
    <p class="lineup"><b>Compared:</b> {chips}</p>
    {PF}
  </div>
  <div class="glance wrap" aria-label="At a glance: {esc(name)}">
    <h3>At a glance</h3>
    <table class="gtable">
      <colgroup><col style="width:12%"><col style="width:29%"><col style="width:25%"><col style="width:34%"></colgroup>
      <thead><tr><th scope="col">Tool</th><th scope="col">Choose this if you are</th><th scope="col">Starting price</th><th scope="col">Main benefit</th></tr></thead>
      <tbody>{grows}</tbody>
    </table>
    <p class="gcap">Full detail for every tool follows on its own page.</p>
    {PF}
  </div>
  <div class="wrap">{profiles}</div>
</section>""")

    contents_rows[group].append((ci, name, cslug))
    chapter_nav.append((ci, group, name, cslug, list(tool_anchors)))
    jobs_html += (f'<li><button type="button" data-job="{cslug}" '
                  f'aria-pressed="false">{esc(job)}</button></li>')

_gl = {}
for n, g, nm, cs, tls in chapter_nav:
    _gl.setdefault(g, []).append((n, nm, cs, tls))
navjs = """
(function(){
  var rail = document.querySelector('.siderail');
  var groups = [].slice.call(document.querySelectorAll('.navgrp[data-chapter]'));
  var byChapter = {};
  groups.forEach(function(g){ byChapter[g.getAttribute('data-chapter')] = g; });
  var toolLinks = {};
  [].slice.call(document.querySelectorAll('.navtools a[data-tool]')).forEach(function(a){
    toolLinks[a.getAttribute('data-tool')] = a;
  });

  /* ---- scroll spy: which chapter and which tool are on screen ---- */
  var chapters = [].slice.call(document.querySelectorAll('section.chapter'));
  var profiles = [].slice.call(document.querySelectorAll('article.profile'));
  var curChapter = null, curTool = null;

  function setChapter(cs){
    if (cs === curChapter) return;
    curChapter = cs;
    groups.forEach(function(g){
      var on = g.getAttribute('data-chapter') === cs;
      g.classList.toggle('is-active', on);
      var d = g.querySelector('details');
      if (d) { if (on) { d.open = true; } else { d.open = false; } }
    });
    var g = byChapter[cs];
    if (g && rail) {
      var gt = g.offsetTop, rt = rail.scrollTop, rh = rail.clientHeight;
      if (gt < rt || gt > rt + rh - 90) rail.scrollTo({top: Math.max(0, gt - 120), behavior: 'smooth'});
    }
  }
  function setTool(id){
    if (id === curTool) return;
    if (curTool && toolLinks[curTool]) toolLinks[curTool].classList.remove('is-here');
    curTool = id;
    if (id && toolLinks[id]) toolLinks[id].classList.add('is-here');
  }

  if ('IntersectionObserver' in window) {
    var chObs = new IntersectionObserver(function(es){
      es.forEach(function(e){ if (e.isIntersecting) setChapter(e.target.id); });
    }, {rootMargin: '-88px 0px -70% 0px', threshold: 0});
    chapters.forEach(function(c){ chObs.observe(c); });

    var tlObs = new IntersectionObserver(function(es){
      es.forEach(function(e){ if (e.isIntersecting) setTool(e.target.id); });
    }, {rootMargin: '-88px 0px -75% 0px', threshold: 0});
    profiles.forEach(function(p){ tlObs.observe(p); });
  }

  /* ---- reading progress + back-to-top ---- */
  var bar = document.getElementById('progress');
  var ticking = false;
  function onScroll(){
    if (ticking) return;
    ticking = true;
    requestAnimationFrame(function(){
      var h = document.documentElement;
      var max = h.scrollHeight - h.clientHeight;
      var pct = max > 0 ? (h.scrollTop / max) * 100 : 0;
      if (bar) bar.style.width = pct.toFixed(2) + '%';
      ticking = false;
    });
  }
  window.addEventListener('scroll', onScroll, {passive: true});
  onScroll();

  /* ---- find-your-tool wizard ---- */
  var WIZ = __WIZDATA__;
  var w1 = document.getElementById('wstep1'), w2 = document.getElementById('wstep2');
  var wq = document.getElementById('wq'), wlist = document.getElementById('wlist');
  var wch = document.getElementById('wchlink');
  if (w1 && w2) {
    [].slice.call(document.querySelectorAll('.wgrid button[data-ch]')).forEach(function(btn){
      btn.addEventListener('click', function(){
        var ch = null;
        for (var i = 0; i < WIZ.length; i++) if (WIZ[i].s === btn.dataset.ch) ch = WIZ[i];
        if (!ch) return;
        wq.textContent = 'Which of these sounds most like you?';
        wlist.innerHTML = '';
        ch.tools.forEach(function(t){
          var b = document.createElement('button');
          b.type = 'button';
          b.innerHTML = '<span class="wp"></span><span class="wt"></span>';
          b.querySelector('.wp').textContent = t[2];
          b.querySelector('.wt').textContent = t[0];
          b.addEventListener('click', function(){ location.hash = t[1]; });
          wlist.appendChild(b);
        });
        wch.setAttribute('href', '#' + ch.s);
        w1.hidden = true; w2.hidden = false;
      });
    });
    var wb = document.getElementById('wback');
    if (wb) wb.addEventListener('click', function(){
      w2.hidden = true; w1.hidden = false;
      wq.textContent = 'What are you looking for?';
    });
  }
})();
"""

nav_links = ('<li class="navsec">Start here</li>'
             '<li class="navgrp"><a class="navflat" href="#contents">Contents</a></li>')
for _g in ("Selling", "Building", "Operating"):
    nav_links += f'<li class="navsec">{esc(_g)}</li>'
    for n, nm, cs, tls in _gl.get(_g, []):
        nav_links += (
            f'<li class="navgrp" data-chapter="{cs}"><details><summary>'
            f'<span class="cn">{n:02d}</span><span class="ct">{esc(nm)}</span></summary>'
            f'<div class="navtools"><a class="navall" href="#{cs}">Chapter overview</a>'
            + "".join(f'<a href="#{a}" data-tool="{a}">{esc(t)}</a>' for t, a in tls)
            + '</div></details></li>')
wizard_buttons = "".join(
    f'<button type="button" data-ch="{c["s"]}"><b>{esc(c["job"])}</b><span>{esc(c["name"])}</span></button>'
    for c in WIZ)
import json as _json
wizard_json = _json.dumps([{ "s": c["s"], "q": c["job"], "tools": c["tools"] } for c in WIZ])

contents_html = ""
for grp in ("Selling", "Building", "Operating"):
    trs = "".join(f'<tr><td class="cnum num">{n:02d}</td>'
                  f'<td><a href="#{s}">{esc(nm)}</a></td></tr>'
                  for n, nm, s in contents_rows[grp])
    contents_html += (f'<div class="grp"><span class="eyebrow">{grp}</span>'
                      f'<table>{trs}</table></div>')

js = """
(function(){
  var search=document.getElementById('toolsearch');
  var results=document.getElementById('results');
  var clearBtn=document.getElementById('clearfilter');
  var state=document.getElementById('filterstate');
  var back=document.getElementById('backfinder');
  var profiles=[].slice.call(document.querySelectorAll('.profile'));
  var rows=[].slice.call(document.querySelectorAll('.gtable tbody tr'));
  var chips=[].slice.call(document.querySelectorAll('.jobs button[data-job]'));

  // index every tool and every chapter so search can jump straight to one
  var index=[];
  document.querySelectorAll('.chapter').forEach(function(ch){
    var title=ch.querySelector('h2').textContent.trim();
    index.push({id:ch.id,label:title,ctx:'Chapter',kind:'chapter',
                key:(title+' '+(ch.querySelector('.decision')||{textContent:''}).textContent).toLowerCase()});
  });
  profiles.forEach(function(p){
    var name=p.querySelector('h4').textContent.trim();
    var chap=p.closest('.chapter').querySelector('h2').textContent.trim();
    index.push({id:p.id,label:name,ctx:chap,kind:'tool',
                key:(name+' '+p.dataset.cat).toLowerCase()});
  });

  var matches=[],active=-1;

  function goTo(id){
    var el=document.getElementById(id);
    if(!el)return;
    closeResults();
    applyFilter('');                       // never leave the reader inside a filtered view
    search.value='';
    if(history.replaceState)history.replaceState(null,'','#'+id);
    el.scrollIntoView({behavior:(window.matchMedia('(prefers-reduced-motion: reduce)').matches?'auto':'smooth'),block:'start'});
    el.setAttribute('tabindex','-1'); el.focus({preventScroll:true});
  }
  function closeResults(){
    results.className='results'; results.innerHTML=''; active=-1; matches=[];
    search.setAttribute('aria-expanded','false');
  }
  function renderResults(q){
    if(!q){closeResults();return;}
    matches=index.filter(function(it){return it.key.indexOf(q)>-1;})
                 .sort(function(a,b){
                    var ai=a.label.toLowerCase().indexOf(q), bi=b.label.toLowerCase().indexOf(q);
                    if((ai===0)!==(bi===0))return ai===0?-1:1;
                    if((ai>-1)!==(bi>-1))return ai>-1?-1:1;
                    return a.kind===b.kind?0:(a.kind==='tool'?-1:1);})
                 .slice(0,8);
    if(!matches.length){closeResults();return;}
    results.innerHTML=matches.map(function(m,i){
      return '<li role="option" aria-selected="false"><button type="button" data-go="'+m.id+'" data-i="'+i+'">'
           + m.label + '<span class="ctx">' + (m.kind==='tool'?m.ctx:'Chapter') + '</span></button></li>';}).join('');
    results.className='results on';
    search.setAttribute('aria-expanded','true');
    active=0; mark();
  }
  function mark(){
    [].slice.call(results.querySelectorAll('button')).forEach(function(b,i){
      b.classList.toggle('active',i===active);
      b.parentNode.setAttribute('aria-selected',i===active?'true':'false');
      if(i===active)b.scrollIntoView({block:'nearest'});
    });
  }
  function applyFilter(q){
    var n=0;
    profiles.forEach(function(p){
      var hit=!q||p.dataset.name.indexOf(q)>-1||p.dataset.cat.indexOf(q)>-1;
      p.classList.toggle('hidden-by-filter',!hit); if(hit)n++;
    });
    rows.forEach(function(r){
      var nm=r.querySelector('.tname').textContent.toLowerCase();
      r.classList.toggle('dim-by-filter',!!q&&nm.indexOf(q)===-1);
    });
    if(q){state.textContent=n+' of '+profiles.length+' tools match \u201c'+q+'\u201d. Press Enter to jump.';
      clearBtn.style.display='inline-block';}
    else{state.textContent='';clearBtn.style.display='none';}
  }
  chips.forEach(function(b){
    b.addEventListener('click',function(){
      chips.forEach(function(o){if(o!==b)o.setAttribute('aria-pressed','false');});
      var on=b.getAttribute('aria-pressed')!=='true';
      b.setAttribute('aria-pressed',on?'true':'false');
      if(on)goTo(b.dataset.job);
      else if(history.replaceState)history.replaceState(null,'',location.pathname);
    });
  });
  if(search){
    search.addEventListener('input',function(){
      var q=search.value.trim().toLowerCase();
      applyFilter(q); renderResults(q);
    });
    search.addEventListener('keydown',function(e){
      if(e.key==='ArrowDown'&&matches.length){e.preventDefault();active=(active+1)%matches.length;mark();}
      else if(e.key==='ArrowUp'&&matches.length){e.preventDefault();active=(active-1+matches.length)%matches.length;mark();}
      else if(e.key==='Enter'&&matches.length){e.preventDefault();goTo(matches[Math.max(active,0)].id);}
      else if(e.key==='Escape'){search.value='';applyFilter('');closeResults();}
    });
    search.addEventListener('blur',function(){setTimeout(closeResults,150);});
  }
  if(results)results.addEventListener('mousedown',function(e){
    var b=e.target.closest('button[data-go]');
    if(b){e.preventDefault();goTo(b.dataset.go);}
  });
  if(clearBtn)clearBtn.addEventListener('click',function(){search.value='';applyFilter('');closeResults();search.focus();});
  if(location.hash.indexOf('#q=')===0){var q=decodeURIComponent(location.hash.slice(3));search.value=q;applyFilter(q.toLowerCase());}
  window.addEventListener('scroll',function(){back.classList.toggle('on',window.scrollY>600);},{passive:true});
})();
"""

doc = f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>The Startup Tool Stack: a field guide to 84 tools in 14 categories</title>
<style>{css}</style>
</head>
<body>
<a class="skip" href="#crm-lead-gen">Skip to chapter 1</a>

<div class="progress" id="progress" aria-hidden="true"></div>
<aside class="siderail" aria-label="Document navigation">
  <a class="rail-home" href="#contents">The Startup Tool Stack</a>
  <ul class="railnav">{nav_links}</ul>
</aside>

<div class="topbar">
  <a class="tb-home" href="#contents">The Startup Tool Stack</a>
  <details class="jump navmenu"><summary>Chapters</summary>
    <ul>{nav_links}</ul>
  </details>
</div>

<header class="mast wrap" id="finder">
  <p class="eyebrow">Tool selection for pre-seed and seed founders</p>
  <h1>The Startup Tool&nbsp;Stack</h1>
  <p class="byline">Last updated August 2026</p>

  <div class="wizard" id="wizard">
    <div class="wlabel">Find your tool</div>
    <h2 class="wq" id="wq">What are you looking for?</h2>
    <div class="wstep" id="wstep1">
      <div class="wgrid">{wizard_buttons}</div>
      <a class="wskip" href="#contents">I'm not sure yet &mdash; show me the full stack</a>
    </div>
    <div class="wstep" id="wstep2" hidden>
      <button class="wback" id="wback" type="button">&larr; All categories</button>
      <div class="wlist" id="wlist"></div>
      <a class="wch" id="wchlink" href="#contents">Or browse this whole chapter</a>
    </div>
  </div>

  <nav class="contents mastcontents" id="contents" aria-label="Contents">
    <h2>Contents</h2>
    {contents_html}
  </nav>
  {PF}
</header>

<section class="howto wrap" id="manual-diy">
  <h2>Why every chapter ends with an internal build</h2>
  <p>Every chapter closes with a seventh option: an internal build, standing the capability up yourself instead of buying it. With AI tools like Claude and Codex, a founder can now get a working version of almost anything in this guide.</p>
  <p>An internal build covers a wide range. It can be a spreadsheet you keep by hand, a script against an API, a self-hosted open-source app, or a full custom build with an AI agent writing most of the code.</p>
  <p>We include it in every comparison because the real question is no longer whether you can build it. It is whether the result is reliable, what it costs you in setup and upkeep, and when that trade beats paying for a product. The internal build pages weigh exactly that.</p>
  {PF}
</section>


<main>{''.join(chapter_html)}</main>


<script>{navjs.replace('__WIZDATA__', wizard_json)}</script>
</body>
</html>"""

_missed = [RIVAL_EDITS[i][0][:60] for i in range(len(RIVAL_EDITS)) if i not in _RIVAL_HITS]
if _missed:
    print("RIVAL_EDITS never matched (%d):" % len(_missed))
    for m in _missed:
        print("  MISS:", m)

with open(OUT, "w") as f:
    f.write(doc)

print(f"written: {OUT}  {len(doc)/1024:.0f} KB")
print(f"chapters: {len(CHAPTERS)} | profiles: {len(all_tool_ids)} ids")
json.dump({"na": NA_LOG, "thin": THIN_LOG, "spans": PAGE_SPANS, "dedup": DEDUP_LOG}, open(f"{HERE}/rollout_log.json", "w"), indent=1)
from collections import Counter as _C
print(f"N/A rewrites: {len(NA_LOG)} | thin cells (<80 chars): {len(THIN_LOG)}")
print("predicted page spans:", dict(_C(n for _c, _t, n in PAGE_SPANS)))
print("leads taken from mid-sentence (setup clause preceded the claim):", len(MIDLEAD_LOG))
for _w, _p, _l in MIDLEAD_LOG[:60]:
    print(f"   {_w}\n      setup kept: \"{_p[:70]}\"\n      lead:       \"{_l[:80]}\"")
print("source-level repeats suppressed (kept in first position):", len(DEDUP_LOG))
for _w, _s in DEDUP_LOG: print(f"   {_w}: {_s}")
for _c, _t, n in PAGE_SPANS:
    if n != 2: print(f"   {n} pages: {_c} / {_t}")
